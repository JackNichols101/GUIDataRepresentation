from typing import Tuple
import PySide2.QtWidgets
import sys
import DemoWindow
import us_state_abbrev
import requests
import Secrets
import sqlite3
from openpyxl import load_workbook

import Map


def get_data():
    all_data = []
    response = requests.get(f"https://api.data.gov/ed/collegescorecard/v1/schools.json?"
                            f"school.degrees_awarded.predominant=2,"
                            f"3&fields=id,school.state,school.name,school.city,2018.student.size,2017.student.size,"
                            f"2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line,"
                            f"2016.repayment.3_yr_repayment.overall&api_key={Secrets.api_key}")
    first_page = response.json()
    if response.status_code != 200:
        print(F"Error Getting Data from API: {response.raw}")
        return []
    total_entries = first_page['metadata']['total']
    entries_per_page = first_page['metadata']['per_page']
    num = (total_entries // entries_per_page) + (total_entries % entries_per_page > 0)
    for page in range(num):
        response = requests.get(
            f"https://api.data.gov/ed/collegescorecard/v1/schools.json?school.degrees_awarded.predominant=2,"
            f"3&fields=id,school.state,school.name,school.city,2018.student.size,2017.student.size,"
            f"2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line,"
            f"2016.repayment.repayment_cohort.3_year_declining_balance,"
            f"2016.repayment.3_yr_repayment.overall&api_key={Secrets.api_key}&page={page}")
        if response.status_code != 200:
            print("error getting data!")
            exit(-1)
        page_of_data = response.json()
        page_of_school_data = page_of_data['results']
        all_data.extend(page_of_school_data)

    return all_data


def open_db(filename: str) -> Tuple[sqlite3.Connection, sqlite3.Cursor]:
    connection = sqlite3.connect(filename)
    cursor = connection.cursor()
    return connection, cursor


def fill_db(all_data, cursor: sqlite3.Cursor):
    query2 = "INSERT INTO colleges VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)"
    for item in all_data:
        if 'PW' != item['school.state'] and 'AS' != item['school.state'] and 'MH' != item['school.state'] \
                and 'FM' != item['school.state'] and 'MP' != item['school.state']:
            cursor.execute(query2, (item['id'], item['school.name'], item['school.city'], item['school.state'],
                                    item['2018.student.size'], item['2017.student.size'],
                                    item['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
                                    item['2016.repayment.repayment_cohort.3_year_declining_balance'],
                                    item['2016.repayment.3_yr_repayment.overall']))


def close_db(connection: sqlite3.Connection):
    connection.commit()
    connection.close()


def setup_db(cursor: sqlite3.Cursor):
    cursor.execute("DROP TABLE IF EXISTS colleges")
    query1 = """CREATE TABLE IF NOT EXISTS colleges(school_id INTEGER PRIMARY KEY, school_name TEXT,
    school_city TEXT, school_state TEXT, student_size_2018 INTEGER, student_size_2017
    INTEGER, earnings_3_yrs_after_completion_overall_count_over_poverty_line_2017
    INTEGER, repayment_cohort_3_year_declining_balance_2016 REAL, repayment_3_yr_repayment_overall_2016 INTEGER)"""
    cursor.execute(query1)


def setup_xl(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    return sheet


def get_xl_data(sheet):
    state_data = []
    max_row = sheet.max_row
    for i in range(1, max_row + 1):
        cell_obj = sheet.cell(row=i, column=10)
        if cell_obj.value == "major":
            state = sheet.cell(row=i, column=2)
            title = sheet.cell(row=i, column=9)
            total_em = sheet.cell(row=i, column=11)
            hourly_25 = sheet.cell(row=i, column=19)
            annually_25 = sheet.cell(row=i, column=24)
            occ_code = sheet.cell(row=i, column=8)
            state_data.append({'state': state.value,
                               'occupation_title': title.value, 'total_employment': total_em.value,
                               'hourly_25': hourly_25.value, 'annually_25': annually_25.value,
                               'occupation_code': occ_code.value})

    return state_data


def setup_db_xl(cursor: sqlite3.Cursor):
    cursor.execute("DROP TABLE IF EXISTS employment")
    query1 = """CREATE TABLE IF NOT EXISTS employment(state TEXT,
    occupation_title TEXT, total_employment INTEGER, hourly_25 INTEGER, annually_25
    INTEGER, occupation_code TEXT)"""
    cursor.execute(query1)


def fill_db_xl(em_data, cursor: sqlite3.Cursor):
    query2 = "INSERT INTO employment VALUES(?, ?, ?, ?, ?, ?)"
    for item in em_data:
        cursor.execute(query2, (item['state'], item['occupation_title'], item['total_employment'], item['hourly_25'],
                                item['annually_25'],
                                item['occupation_code']))


def start_widget(data, html):
    qt_app = PySide2.QtWidgets.QApplication(sys.argv)
    my_window = DemoWindow.Comp490DemoWindow(data, html)
    sys.exit(qt_app.exec_())


def update_file(file, cursor):
    sheet = setup_xl(file)
    state_data = get_xl_data(sheet)
    setup_db_xl(cursor)
    fill_db_xl(state_data, cursor)
    print(state_data)
    data, html = organize_data_for_widget(cursor)
    print(data)
    return data, html


def organize_data_for_widget(cursor):
    query1 = """SELECT school_state, SUM(student_size_2018) as tot, AVG(repayment_cohort_3_year_declining_balance_2016)
    FROM colleges
    GROUP BY school_state"""
    query2 = """SELECT state, SUM(total_employment) as em, CAST(AVG(annually_25) AS Integer)
    FROM employment
    WHERE CAST(SUBSTR(occupation_code, 1, 2) AS integer) NOT BETWEEN 30 AND 50
    GROUP BY state"""
    cursor.execute(query1)
    one = list(cursor.fetchall())
    print(one)
    cursor.execute(query2)
    two = list(cursor.fetchall())
    data = {}
    for item in one:
        grad = item[1]
        grad = int(grad / 4)
        data[us_state_abbrev.abbrev_us_state[item[0]]] = [grad, item[2], 0, 0]
    for item in two:
        if item[0] in data:
            data[item[0]][2] = item[1]  # data[item[0]].append(item[1])
            data[item[0]][3] = item[2]  # data[item[0]].append(item[2])
    print(data)
    html_map = Map.convert(data)
    dat = []
    for key, value in data.items():
        it = [key, value[0], value[2], value[3], value[1]]
        dat.append(it)
    print(dat)
    return dat, html_map


def main():
    demo_data = get_data()
    sheet = setup_xl("state_M2019_dl.xlsx")
    connection, cursor = open_db('college_data.db')
    state_data = get_xl_data(sheet)
    setup_db_xl(cursor)
    fill_db_xl(state_data, cursor)
    setup_db(cursor)
    fill_db(demo_data, cursor)
    data, html = organize_data_for_widget(cursor)
    close_db(connection)
    start_widget(data, html)


if __name__ == '__main__':
    main()
